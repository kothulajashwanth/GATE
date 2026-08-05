"""Report generation: PDF, Excel, CSV for exam results."""

from io import BytesIO

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph


def _rows_for_results(results: list) -> list[list]:
    rows = [["Roll", "Name", "Department", "Semester", "Score", "Marks", "Percentage", "Rank", "Pass"]]
    for r in results:
        student = r.student
        user = student.user if student else None
        rows.append([
            student.roll_number if student else "",
            user.full_name if user else "",
            student.department.name if student and student.department else "",
            student.semester.name if student and student.semester else "",
            f"{r.obtained_marks}",
            f"{r.total_marks}",
            f"{r.percentage or 0:.1f}%",
            str(r.rank or ""),
            "Yes" if r.is_passed else "No",
        ])
    return rows


def generate_pdf(exam_title: str, results: list) -> BytesIO:
    """PDF result sheet for an exam."""
    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf,
        pagesize=landscape(A4),
        rightMargin=30, leftMargin=30, topMargin=30, bottomMargin=30,
    )
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle("Title2", parent=styles["Title"], fontSize=18, spaceAfter=12)
    elems = [Paragraph(f"Result Report: {exam_title}", title_style)]

    rows = _rows_for_results(results)
    if rows:
        table = Table(rows, repeatRows=1)
        table.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1e40af")),
            ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
            ("FONTSIZE", (0, 0), (-1, -1), 8),
            ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#f1f5f9")]),
            ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
        ]))
        elems.append(table)
    else:
        elems.append(Paragraph("No results available.", styles["BodyText"]))

    doc.build(elems)
    buf.seek(0)
    return buf


def generate_excel(exam_title: str, results: list) -> BytesIO:
    """Excel result sheet."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()
    ws = wb.active
    ws.title = "Results"
    header = ["Roll", "Name", "Department", "Semester", "Score", "Marks", "Percentage", "Rank", "Pass"]
    ws.append([f"Result Report: {exam_title}"])
    ws.append([])
    ws.append(header)
    for cell in ws[3]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="1E40AF", end_color="1E40AF", fill_type="solid")

    for r in results:
        student = r.student
        user = student.user if student else None
        ws.append([
            student.roll_number if student else "",
            user.full_name if user else "",
            student.department.name if student and student.department else "",
            student.semester.name if student and student.semester else "",
            r.obtained_marks,
            r.total_marks,
            r.percentage or 0,
            r.rank or "",
            "Yes" if r.is_passed else "No",
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def generate_csv(exam_title: str, results: list) -> BytesIO:
    """CSV result sheet (UTF-8 BOM for Excel compatibility)."""
    import csv
    import io

    buf = BytesIO()
    wrapper = io.TextIOWrapper(buf, encoding="utf-8-sig")
    writer = csv.writer(wrapper)
    writer.writerow([f"Result Report: {exam_title}"])
    writer.writerow(["Roll", "Name", "Department", "Semester", "Score", "Marks", "Percentage", "Rank", "Pass"])
    for r in results:
        student = r.student
        user = student.user if student else None
        writer.writerow([
            student.roll_number if student else "",
            user.full_name if user else "",
            student.department.name if student and student.department else "",
            student.semester.name if student and student.semester else "",
            r.obtained_marks,
            r.total_marks,
            r.percentage or 0,
            r.rank or "",
            "Yes" if r.is_passed else "No",
        ])
    wrapper.detach()
    buf.seek(0)
    return buf