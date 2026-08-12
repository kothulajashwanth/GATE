from dataclasses import dataclass, field
from io import BytesIO
from typing import Any

from openpyxl import Workbook, load_workbook
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.errors import ValidationError
from app.db.models.academic import Department, Section, Semester
from app.db.models.student import Student
from app.db.models.user import Role, User

EXPECTED_HEADERS = [
    "full_name",
    "email",
    "roll_number",
    "phone",
    "department",
    "semester",
    "section",
    "status",
]


@dataclass
class ImportResult:
    total: int = 0
    imported: int = 0
    errors: list[dict[str, Any]] = field(default_factory=list)

    @property
    def failed(self) -> int:
        return len(self.errors)


def generate_student_template() -> BytesIO:
    """Generate an Excel template (.xlsx) for student bulk import."""
    wb = Workbook()
    ws = wb.active
    ws.title = "Student Roster Template"
    ws.append(EXPECTED_HEADERS)

    # Sample rows for user guidance
    ws.append(["Alex Johnson", "alex.johnson@college.edu", "CS2024001", "9876543210", "CSE", "1", "A", "ACTIVE"])
    ws.append(["Sarah Smith", "sarah.smith@college.edu", "ECE2024002", "9876543211", "ECE", "2", "B", "ACTIVE"])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def _row_error(row_number: int, message: str) -> dict[str, Any]:
    return {"row": row_number, "error": message}


async def validate_student_import(db: AsyncSession, file: BytesIO) -> dict[str, Any]:
    """Parse and validate Excel workbook without inserting. Return validation summary for preview UI."""
    try:
        wb = load_workbook(file, read_only=True, data_only=True)
    except Exception:
        raise ValidationError("Could not parse the Excel file. Use the provided .xlsx template.")

    ws = wb.active
    rows = ws.iter_rows(values_only=True)
    try:
        header_row = next(rows)
        header = [str(c).strip().lower() if c else "" for c in header_row]
    except StopIteration:
        raise ValidationError("The Excel file is empty")

    # Flexible header checking (supports full_name/name, roll_number/rollNumber, etc.)
    header_str = ",".join(header)
    if "email" not in header_str or ("roll" not in header_str and "roll_number" not in header_str):
        raise ValidationError(
            f"Invalid file headers. Required headers include: {', '.join(EXPECTED_HEADERS)}"
        )

    # Reference mappings from DB
    dept_rows = (await db.execute(select(Department))).scalars().all()
    dept_map: dict[str, str] = {}
    for d in dept_rows:
        dept_map[d.code.upper()] = str(d.id)
        dept_map[d.name.upper()] = str(d.id)

    sem_rows = (await db.execute(select(Semester))).scalars().all()
    sem_map: dict[tuple[str, str], str] = {}
    for s in sem_rows:
        sem_map[(str(s.department_id), str(s.ordinal))] = str(s.id)
        sem_map[(str(s.department_id), s.name.upper())] = str(s.id)

    sec_rows = (await db.execute(select(Section))).scalars().all()
    sec_map: dict[tuple[str, str], str] = {}
    for sc in sec_rows:
        sec_map[(str(sc.semester_id), sc.code.upper())] = str(sc.id)
        sec_map[(str(sc.semester_id), sc.name.upper())] = str(sc.id)

    existing_rolls = {r[0] for r in (await db.execute(select(Student.roll_number))).all()}
    existing_emails = {r[0] for r in (await db.execute(select(User.email))).all()}

    valid_rows: list[dict[str, Any]] = []
    errors: list[dict[str, Any]] = []
    seen_rolls: set[str] = set()
    seen_emails: set[str] = set()

    total = 0
    row_number = 1
    for raw in rows:
        row_number += 1
        if raw is None or all(c is None or str(c).strip() == "" for c in raw):
            continue
        total += 1

        def cell(idx: int) -> str:
            v = raw[idx] if idx < len(raw) else None
            return str(v).strip() if v is not None else ""

        full_name = cell(0)
        email = cell(1).lower()
        roll = cell(2)
        phone = cell(3)
        dept_str = cell(4).upper()
        sem_str = cell(5).upper()
        sec_str = cell(6).upper()
        status_str = cell(7).upper() or "ACTIVE"

        row_errors: list[str] = []
        if not full_name:
            row_errors.append("Full name is required")
        if not email:
            row_errors.append("Email is required")
        elif "@" not in email:
            row_errors.append(f"Invalid email format '{email}'")
        elif email in existing_emails or email in seen_emails:
            row_errors.append(f"Duplicate email '{email}'")

        if not roll:
            row_errors.append("Roll number is required")
        elif roll in existing_rolls or roll in seen_rolls:
            row_errors.append(f"Duplicate roll number '{roll}'")

        dept_id = dept_map.get(dept_str)
        if dept_str and not dept_id:
            row_errors.append(f"Unknown department '{dept_str}'")

        sem_id = sem_map.get((dept_id, sem_str)) if dept_id else None
        if sem_str and dept_id and not sem_id:
            row_errors.append(f"Unknown semester '{sem_str}' for department '{dept_str}'")

        sec_id = sec_map.get((sem_id, sec_str)) if sem_id else None
        if sec_str and sem_id and not sec_id:
            row_errors.append(f"Unknown section '{sec_str}' for semester '{sem_str}'")

        if row_errors:
            errors.append(_row_error(row_number, "; ".join(row_errors)))
        else:
            seen_rolls.add(roll)
            seen_emails.add(email)
            name_parts = full_name.split(" ", 1)
            first_name = name_parts[0]
            last_name = name_parts[1] if len(name_parts) > 1 else ""

            valid_rows.append({
                "row": row_number,
                "roll_number": roll,
                "full_name": full_name,
                "first_name": first_name,
                "last_name": last_name,
                "email": email,
                "phone": phone or None,
                "department_id": dept_id,
                "semester_id": sem_id,
                "section_id": sec_id,
                "status": status_str,
            })

    return {
        "total": total,
        "valid_count": len(valid_rows),
        "invalid_count": len(errors),
        "valid_rows": valid_rows,
        "errors": errors,
    }


async def import_students_from_excel(db: AsyncSession, file: BytesIO) -> ImportResult:
    """Validate and commit student roster to PostgreSQL."""
    validation = await validate_student_import(db, file)
    result = ImportResult(total=validation["total"], errors=validation["errors"])
    if validation["errors"]:
        return result

    for p in validation["valid_rows"]:
        user = User(
            email=p["email"],
            first_name=p["first_name"],
            last_name=p["last_name"] or None,
            phone=p["phone"],
            role=Role.STUDENT,
            is_active=(p["status"] != "INACTIVE"),
        )
        db.add(user)
        await db.flush()

        student = Student(
            user_id=user.id,
            roll_number=p["roll_number"],
            department_id=p["department_id"],
            semester_id=p["semester_id"],
            section_id=p["section_id"],
            phone=p["phone"],
        )
        db.add(student)

    await db.commit()
    result.imported = len(validation["valid_rows"])
    return result